"""
Generate the invoice xlsx for the Cornbread Hemp × Abbey Road on the River proposal.
Producer: AGV Miami (legal entity AGV Miami, LLC).
Layout mirrors the Gymshark Phone Box invoice.
Run: python3 scripts/generate_invoice.py
Output: public/invoices/CBH-ABR-2026-V2.1.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = "public/invoices/CBH-ABR-2026-V2.1.xlsx"

# Cornbread palette tuned to match the proposal page (orange + cedar)
ORANGE = "C5883F"
ORANGE_DARK = "8E5C24"
DARK = "1A140E"
INK = "2A211A"
MUTED = "8A8580"
WHITE = "FFFFFF"
BUTTERMILK = "F5EFDF"

thin = Side(style="thin", color="DDDDDD")
border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

f_title = Font(name="Helvetica", size=22, bold=True, color=DARK)
f_eyebrow = Font(name="Helvetica", size=9, bold=True, color=ORANGE)
f_label = Font(name="Helvetica", size=9, bold=True, color=MUTED)
f_meta = Font(name="Helvetica", size=11, color=DARK)
f_phase = Font(name="Helvetica", size=11, bold=True, color=WHITE)
f_line_name = Font(name="Helvetica", size=10, bold=True, color=DARK)
f_line_desc = Font(name="Helvetica", size=9, color=MUTED)
f_amount = Font(name="Helvetica", size=10, bold=True, color=DARK)
f_basis = Font(name="Helvetica", size=8, color=MUTED, italic=True)
f_subtotal = Font(name="Helvetica", size=10, bold=True, color=ORANGE_DARK)
f_total = Font(name="Helvetica", size=14, bold=True, color=DARK)
f_total_amount = Font(name="Helvetica", size=14, bold=True, color=ORANGE_DARK)
f_footer = Font(name="Helvetica", size=8, color=MUTED, italic=True)

fill_phase = PatternFill("solid", start_color=INK, end_color=INK)
fill_subtotal = PatternFill("solid", start_color=BUTTERMILK, end_color=BUTTERMILK)
fill_total = PatternFill("solid", start_color=ORANGE, end_color=ORANGE)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice CBH-ABR V2.1"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 64
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 4

    row = 1

    # Header band
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="AGV MIAMI")
    cell.font = Font(name="Helvetica", size=24, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="EXPERIENTIAL FABRICATION & PRODUCTION  ·  PROPOSAL INVOICE")
    cell.font = Font(name="Helvetica", size=9, bold=True, color=ORANGE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    row += 1

    # Two-column header: Producer (left) vs Client/Project (right)
    ws.cell(row=row, column=2, value="PRODUCER").font = f_label
    ws.cell(row=row, column=3, value="CLIENT").font = f_label
    ws.cell(row=row, column=4, value="PROJECT").font = f_label
    row += 1

    ws.cell(row=row, column=2, value="AGV Miami, LLC").font = f_meta
    ws.cell(row=row, column=3, value="Cornbread Hemp").font = f_meta
    ws.cell(row=row, column=4, value="CBH-ABR-2026").font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="Miami, FL").font = f_meta
    ws.cell(row=row, column=3, value="Attn: Jess Gago").font = f_meta
    ws.cell(row=row, column=4, value="Version 2.1").font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="jclarkson@agvmiami.com").font = f_meta
    ws.cell(row=row, column=3, value="Brand Activations Manager").font = f_meta
    ws.cell(row=row, column=4, value="Issue Date: May 5, 2026").font = f_meta
    row += 1

    ws.cell(row=row, column=4, value="Activation: May 21–25, 2026").font = f_meta
    row += 1

    ws.cell(row=row, column=4, value="Venue: RiverStage, Jeffersonville, IN").font = f_meta
    row += 2

    # Project description
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="ENGAGEMENT")
    cell.font = f_eyebrow
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=(
        "Cornbread Hemp's existing activation is picked up from the holding vendor in Louisville, KY on Tuesday, May 19, "
        "transported to RiverStage in Jeffersonville, IN, and installed by end of day so the activation is VIP-ready by sunset on Wednesday, May 20. "
        "Five public event days follow. Strike on Tuesday, May 26 with long-haul return to the AGV Miami warehouse for ongoing storage and 48-hour mobilization to the next event on the calendar."
    ))
    cell.font = Font(name="Helvetica", size=10, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    row += 2

    # Pricing rate card section
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="PRICING — PUBLISHED RATE CARD")
    cell.font = f_eyebrow
    row += 1

    def section_header(label, eyebrow):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value=f"{label}   ·   {eyebrow}")
        c.font = f_phase
        c.fill = fill_phase
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

    def line_item(name, desc, amount, basis, height=44, starting_at=False, amount_text=None):
        nonlocal row
        ws.cell(row=row, column=2, value=name).font = f_line_name
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=3, value=desc).font = f_line_desc
        ws.cell(row=row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        if amount_text is not None:
            amount_label = amount_text
        else:
            amount_label = f"{'from ' if starting_at else ''}${amount:,}"
        amount_cell = ws.cell(row=row, column=4)
        amount_cell.value = f"{amount_label}\n{basis}"
        amount_cell.font = f_amount
        amount_cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = height
        row += 1

    # Per Activation
    section_header("Per Activation", "Costs that move with each event")
    line_item(
        "Build & Strike — Logistics",
        "Truck rental, fuel, freight insurance, and round-trip transport from the AGV Miami warehouse to the venue.",
        4500,
        "per activation",
    )
    line_item(
        "Build & Strike — Labor (Local Hires)",
        "Two-crew lift sourced near the venue for build and strike days, supervised on site by the AGV producer.",
        1800,
        "per activation",
    )
    line_item(
        "Build & Strike — Travel & Lodging (1–2 AGV Crew)",
        "Round-trip travel and lodging for the AGV producer (and second lead on long-haul routes). Itemized at cost.",
        2500,
        "per activation",
    )

    # Per Month
    section_header("Per Month", "Asset care between events")
    line_item(
        "Storage — Monthly Hold (NYC or Miami)",
        "Climate-monitored, alarmed warehouse hold at the AGV facility of choice. Includes inventory tracking, condition logging, and 48-hour mobilization for the next event.",
        1500,
        "per month",
        height=52,
    )

    # Upgrades
    section_header("Upgrades", "Investments in the activation itself")
    line_item(
        "Plants — Foliage Rental (Per Activation Instance)",
        "Commercial-grade outdoor artificial foliage scaled to the planter ring. Sourced, planted, and pulled each instance — same kit re-deploys cleanly across the calendar.",
        1200,
        "per instance",
        height=52,
    )
    line_item(
        "Lighting — Upgrade + Routine Maintenance",
        "One-time fixture purchase tuned for evening illumination, with astronomical timer programming and a single venue drop. Routine maintenance carries forward across deployments.",
        4500,
        "one-time",
        height=52,
    )

    # Optional Add-Ons
    section_header("Optional Add-Ons", "Pre-priced overlays, available on request")
    line_item(
        "Pre-Deployment Refresh",
        "Touch-up pass: paint, hardware, finishes, lighting recalibration. Scaled to the asset's condition on intake.",
        2500,
        "per instance",
        starting_at=True,
    )
    line_item(
        "Pre-Deployment Rebrand",
        "Graphics, wraps, or finish refresh aligned with a new campaign or partnership. Designed at the front end.",
        7500,
        "per instance",
        starting_at=True,
    )
    line_item(
        "On-Site Show-Day Coverage",
        "AGV producer or specialized tech on the ground during show days — refresh, repair, or in-window punch-list response.",
        850,
        "per day",
        starting_at=True,
    )

    # Change Orders — commercial mechanics
    section_header("Change Orders", "Commercial mechanics — bundle savings, pass-through, overflow storage")
    line_item(
        "Calendar Bundle Discount",
        "4+ events committed inside a single 12-month window unlock 6–10% off per-activation logistics. Stacks on the rate card.",
        0,
        "stacks on rate card",
        amount_text="6–10%",
    )
    line_item(
        "Travel Pass-Through (Above Cap)",
        "Travel or lodging beyond the per-activation Travel & Lodging line cap bills at cost with receipts on written authorization.",
        0,
        "with receipts",
        amount_text="At cost",
    )
    line_item(
        "Additional Storage (Overflow Pallets)",
        "Storage beyond the standard footprint — overflow pallets, additional environmental control — bills monthly per pallet.",
        400,
        "per pallet / month",
    )

    row += 1

    # Abbey Road — Applied Composition
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="ABBEY ROAD · APPLIED COMPOSITION")
    cell.font = f_eyebrow
    row += 1

    def applied_line(label, note, amount):
        nonlocal row
        ws.cell(row=row, column=2, value=label).font = f_line_name
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=3, value=note).font = f_line_desc
        ws.cell(row=row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        c = ws.cell(row=row, column=4, value=amount)
        c.font = f_amount
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    applied_line("Build & Strike — Logistics", "Asymmetric route absorbed inside the baseline.", 4500)
    applied_line("Build & Strike — Labor (Local Hires)", "", 1800)
    applied_line("Build & Strike — Travel & Lodging (1–2 AGV Crew)", "", 2500)
    applied_line("Plants — Foliage Rental", "Per activation instance.", 1200)
    applied_line("Lighting — Upgrade + Routine Maintenance", "Year-1 only — covers the calendar going forward.", 4500)

    # Subtotals row
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = fill_subtotal
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value="Per activation $8,800   ·   Upgrades $5,700")
    c.font = f_subtotal
    c.alignment = Alignment(horizontal="right", indent=1)
    c2 = ws.cell(row=row, column=4, value="")
    ws.row_dimensions[row].height = 22
    row += 1

    # Net total bar
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = fill_total
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value="ABBEY ROAD ENGAGEMENT NET")
    c.font = Font(name="Helvetica", size=14, bold=True, color=DARK)
    c.alignment = Alignment(horizontal="right", indent=1, vertical="center")
    c2 = ws.cell(row=row, column=4, value=14500)
    c2.font = Font(name="Helvetica", size=18, bold=True, color=DARK)
    c2.number_format = '"$"#,##0'
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 32
    row += 2

    # 2026 Calendar — preview
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="2026 CALENDAR — PER-EVENT PROJECTION")
    cell.font = f_eyebrow
    row += 1

    def calendar_row(event, dates_city, miles_market, amount, note=""):
        nonlocal row
        ws.cell(row=row, column=2, value=event).font = f_line_name
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        desc = f"{dates_city}  ·  {miles_market}"
        if note:
            desc += f"  ·  {note}"
        ws.cell(row=row, column=3, value=desc).font = f_line_desc
        ws.cell(row=row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        c = ws.cell(row=row, column=4, value=amount)
        c.font = f_amount
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right", vertical="top")
        ws.row_dimensions[row].height = 24
        row += 1

    calendar_row("Abbey Road on the River", "May 21–25, Jeffersonville, IN", "1,200 mi · out-of-market", 14500, "incl. lighting upgrade")
    calendar_row("Minnesota Yacht Club Festival", "Jul 17–19, St. Paul, MN", "1,800 mi · out-of-market", 10000, "per-activation + plants")
    calendar_row("North Carolina Folk Festival", "Sept 18–20, Greensboro, NC", "800 mi · out-of-market", 10000, "per-activation + plants")
    calendar_row("Fin Fest", "Nov 13–14, St. Petersburg, FL", "270 mi · local market", 6500, "no lodging line, day trips")

    # Annual production
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = fill_subtotal
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value="Annual production — 4 events + year-1 lighting upgrade")
    c.font = f_subtotal
    c.alignment = Alignment(horizontal="right", indent=1)
    c2 = ws.cell(row=row, column=4, value=41000)
    c2.font = f_subtotal
    c2.number_format = '"$"#,##0'
    c2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 22
    row += 1

    # Storage note
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value="+ Storage at $1,500/month rolling — billed monthly between deployments.")
    c.font = f_basis
    c.alignment = Alignment(horizontal="right", indent=1)
    row += 2

    # Payment Terms
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="PAYMENT TERMS")
    cell.font = f_eyebrow
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=(
        "50% deposit ($7,250) due upon Client's written approval of this Scope of Work (Proposal execution). "
        "50% balance ($7,250) due on or before May 12, 2026 (pre-event walkthrough date). "
        "Payment exclusively via ACH electronic transfer or domestic wire transfer; ACH/wire details issued directly to the Client billing contact upon SoW approval. "
        "Reference CBH-ABR-2026 V2.1 on all remittances."
    ))
    cell.font = Font(name="Helvetica", size=10, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 70
    row += 2

    # Cancellation
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="CANCELLATION & CHANGE ORDERS")
    cell.font = f_eyebrow
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=(
        "Full refund if cancelled before May 5. 50% refund between May 5 — May 12. No refund after May 12 (truck, driver, plants, and lighting locked). "
        "Anything added — additional crew, on-site coverage, refresh, rebrand — is quoted as a written change order before execution. Nothing billed that wasn't signed."
    ))
    cell.font = Font(name="Helvetica", size=10, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    row += 2

    # Footer
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=(
        "AGV Miami, LLC  ·  jclarkson@agvmiami.com  ·  This invoice is issued in connection with proposal "
        "CBH-ABR-2026 V2.1 and incorporates by reference the executed Master Services Agreement between AGV Miami, LLC and Cornbread Hemp."
    ))
    cell.font = f_footer
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[row].height = 36

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
